import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import datetime
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook

class ReportExporter:
    def __init__(self, project_name, fluid_type):
        self.project_name = project_name
        self.fluid_type = fluid_type

    def _auto_adjust_columns(self, worksheet):
        """Автоподбор ширины столбцов по содержимому"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_value = str(cell.value)
                        # Учитываем переносы строк
                        lines = cell_value.split('\n')
                        for line in lines:
                            max_length = max(max_length, len(line))
                except:
                    pass
            
            # Добавляем отступ (коэффициент можно настроить)
            adjusted_width = min(max_length + 2, 50)  # Максимальная ширина 50 символов
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _apply_styles(self, worksheet):
        """Применяет стили к листу Excel"""
        # Определяем стили
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Стиль для заголовков
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Стиль для данных
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if cell.row % 2 == 0:  # Чередование цвета строк
                    cell.fill = cell_fill
                
                # Выравнивание
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

    def _get_component_info(self, section):
        """Получает информацию о компонентах сложного участка"""
        components_info = []
        components = section.get("components", [])
        
        for i, comp in enumerate(components):
            comp_info = {
                'Название участка': section.get('name', 'N/A'),
                'Название компонента': comp.get('name', f'Компонент {i+1}'),
                'Тип компонента': comp.get('component_type', 'unknown'),
                'Материал': comp.get('material', 'N/A'),
                'Длина, м': comp.get('length', 0),
                'Диаметр, мм': comp.get('diameter', 0),
                'Толщина стенки, мм': comp.get('thickness', comp.get('wall_thickness', 0)),
                'Количество': comp.get('count', 1),
                'Прокладка': section.get('location', 'надземная'),
                'Защита': section.get('protection', 'без защиты'),
                'Среда': section.get('environment', 'Поволжье')
            }
            
            # Остаточная толщина и состояние
            remaining = comp.get('remaining', comp_info['Толщина стенки, мм'])
            comp_info['Остаточная толщина, мм'] = round(remaining, 2)
            
            # Определяем уровень коррозии
            from models.corrosion import get_corrosion_level
            level, _ = get_corrosion_level(remaining)
            comp_info['Уровень коррозии'] = level
            
            components_info.append(comp_info)
        
        return components_info

    def export_to_excel(self, sections_data, economic_summary, filename=None):
        """Экспорт в Excel с поддержкой сложных участков"""
        try:
            # Если имя файла не указано - генерируем автоматически
            if not filename:
                timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
                filename = f"отчет_трубопровод_{timestamp}.xlsx"
            
            # Проверяем расширение файла
            if not filename.lower().endswith('.xlsx'):
                filename += '.xlsx'
            
            # Создаем новый Excel файл с несколькими листами
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                workbook = writer.book
                
                # ===============================================
                # ЛИСТ 1: Сводная информация по участкам
                # ===============================================
                sections_list = []
                total_components = 0
                total_length = 0
                
                for section in sections_data:
                    components = section.get("components", [])
                    
                    # Для простых участков (для совместимости)
                    if not components:
                        section_info = {
                            'Название': section.get('name', 'N/A'),
                            'Тип': 'простой участок',
                            'Кол-во компонентов': 1,
                            'Длина, м': section.get('length', 0),
                            'Диаметр, мм': section.get('diameter', 0),
                            'Толщина, мм': section.get('thickness', 0),
                            'Материал': section.get('material', 'N/A'),
                            'Прокладка': section.get('location', 'надземная'),
                            'Защита': section.get('protection', 'без защиты'),
                            'Среда': section.get('environment', 'Поволжье'),
                            'Остаточная толщина, мм': round(section.get('remaining_thickness', section.get('thickness', 0)), 2)
                        }
                        
                        # Определяем уровень коррозии
                        from models.corrosion import get_corrosion_level
                        remaining = section_info['Остаточная толщина, мм']
                        level, _ = get_corrosion_level(remaining)
                        section_info['Уровень коррозии'] = level
                        
                        sections_list.append(section_info)
                        total_components += 1
                        total_length += section_info['Длина, м']
                        
                    else:
                        # Для сложных участков
                        total_components += len(components)
                        
                        # Считаем суммарную длину трубных компонентов
                        section_length = 0
                        for comp in components:
                            if comp.get('component_type') == 'pipe':
                                section_length += comp.get('length', 0)
                        
                        total_length += section_length
                        
                        # Определяем худшее состояние среди компонентов
                        worst_remaining = 100  # большое число
                        worst_component = None
                        
                        for comp in components:
                            remaining = comp.get('remaining', 
                                                comp.get('thickness', 
                                                        comp.get('wall_thickness', 10)))
                            if remaining < worst_remaining:
                                worst_remaining = remaining
                                worst_component = comp
                        
                        # Сводная информация по участку
                        section_info = {
                            'Название': section.get('name', 'N/A'),
                            'Тип': section.get('object_type', 'сложный объект'),
                            'Кол-во компонентов': len(components),
                            'Длина, м': section_length,
                            'Диаметр, мм': 'разные' if components else 0,
                            'Толщина, мм': 'разные' if components else 0,
                            'Материал': 'разные' if len(set(c.get('material', '') for c in components)) > 1 else 
                                      (components[0].get('material', 'N/A') if components else 'N/A'),
                            'Прокладка': section.get('location', 'надземная'),
                            'Защита': section.get('protection', 'без защиты'),
                            'Среда': section.get('environment', 'Поволжье'),
                            'Остаточная толщина, мм': round(worst_remaining, 2) if worst_component else 0
                        }
                        
                        # Определяем уровень коррозии по худшему компоненту
                        from models.corrosion import get_corrosion_level
                        level, _ = get_corrosion_level(worst_remaining)
                        section_info['Уровень коррозии'] = level
                        
                        sections_list.append(section_info)
                
                sections_df = pd.DataFrame(sections_list)
                sections_df.to_excel(writer, sheet_name='Участки_сводка', index=False)
                worksheet = writer.sheets['Участки_сводка']
                self._apply_styles(worksheet)
                self._auto_adjust_columns(worksheet)
                
                # ===============================================
                # ЛИСТ 2: Детальная информация по компонентам
                # ===============================================
                all_components = []
                
                for section in sections_data:
                    if section.get("components"):
                        # Сложный участок
                        components_info = self._get_component_info(section)
                        all_components.extend(components_info)
                    else:
                        # Простой участок
                        component_info = {
                            'Название участка': section.get('name', 'N/A'),
                            'Название компонента': section.get('name', 'N/A'),
                            'Тип компонента': 'труба',
                            'Материал': section.get('material', 'N/A'),
                            'Длина, м': section.get('length', 0),
                            'Диаметр, мм': section.get('diameter', 0),
                            'Толщина стенки, мм': section.get('thickness', 0),
                            'Количество': 1,
                            'Прокладка': section.get('location', 'надземная'),
                            'Защита': section.get('protection', 'без защиты'),
                            'Среда': section.get('environment', 'Поволжье')
                        }
                        
                        # Остаточная толщина и состояние
                        remaining = section.get('remaining_thickness', component_info['Толщина стенки, мм'])
                        component_info['Остаточная толщина, мм'] = round(remaining, 2)
                        
                        from models.corrosion import get_corrosion_level
                        level, _ = get_corrosion_level(remaining)
                        component_info['Уровень коррозии'] = level
                        
                        all_components.append(component_info)
                
                if all_components:
                    components_df = pd.DataFrame(all_components)
                    components_df.to_excel(writer, sheet_name='Компоненты', index=False)
                    worksheet = writer.sheets['Компоненты']
                    self._apply_styles(worksheet)
                    self._auto_adjust_columns(worksheet)
                
                # ===============================================
                # ЛИСТ 3: Экономика
                # ===============================================
                economic_data = {
                    'Параметр': [
                        'Всего участков',
                        'Всего компонентов',
                        'Срочный ремонт (аварийные участки)',
                        'Плановый ремонт', 
                        'Текущее обслуживание',
                        'Стоимость срочного ремонта, руб',
                        'Стоимость планового ремонта, руб',
                        'Стоимость обслуживания, руб',
                        'ОБЩАЯ СТОИМОСТЬ, руб'
                    ],
                    'Значение': [
                        len(sections_data),
                        total_components,
                        economic_summary.get('urgent_count', 0),
                        economic_summary.get('planned_count', 0),
                        len(sections_data) - economic_summary.get('urgent_count', 0) - economic_summary.get('planned_count', 0),
                        economic_summary.get('urgent_repair_cost', 0),
                        economic_summary.get('planned_repair_cost', 0),
                        economic_summary.get('maintenance_cost', 0),
                        economic_summary.get('total_repair_cost', 0)
                    ]
                }
                economic_df = pd.DataFrame(economic_data)
                economic_df.to_excel(writer, sheet_name='Экономика', index=False)
                worksheet = writer.sheets['Экономика']
                self._apply_styles(worksheet)
                self._auto_adjust_columns(worksheet)
                
                # ===============================================
                # ЛИСТ 4: Технические параметры
                # ===============================================
                # Рассчитываем средние значения
                avg_diameter = 0
                avg_thickness = 0
                diameter_count = 0
                thickness_count = 0
                
                for section in sections_data:
                    if section.get("components"):
                        for comp in section["components"]:
                            if comp.get('diameter', 0) > 0:
                                avg_diameter += comp['diameter']
                                diameter_count += 1
                            if comp.get('thickness', 0) > 0:
                                avg_thickness += comp['thickness']
                                thickness_count += 1
                            elif comp.get('wall_thickness', 0) > 0:
                                avg_thickness += comp['wall_thickness']
                                thickness_count += 1
                    else:
                        if section.get('diameter', 0) > 0:
                            avg_diameter += section['diameter']
                            diameter_count += 1
                        if section.get('thickness', 0) > 0:
                            avg_thickness += section['thickness']
                            thickness_count += 1
                
                avg_diameter = avg_diameter / diameter_count if diameter_count > 0 else 0
                avg_thickness = avg_thickness / thickness_count if thickness_count > 0 else 0
                
                params_data = {
                    'Параметр': [
                        'Тип среды',
                        'Дата формирования отчета',
                        'Общее количество участков',
                        'Общее количество компонентов',
                        'Общая длина трубопровода, м',
                        'Средний диаметр, мм',
                        'Средняя толщина стенки, мм',
                        'Процент износа системы',
                        'Рекомендуемый бюджет, руб'
                    ],
                    'Значение': [
                        'НЕФТЬ' if self.fluid_type == 'oil' else 'ГАЗ',
                        datetime.datetime.now().strftime('%d.%m.%Y %H:%M'),
                        len(sections_data),
                        total_components,
                        total_length,
                        round(avg_diameter, 1),
                        round(avg_thickness, 2),
                        f"{((economic_summary.get('urgent_count', 0) / max(len(sections_data), 1)) * 100):.1f}%",
                        economic_summary.get('total_repair_cost', 0)
                    ]
                }
                params_df = pd.DataFrame(params_data)
                params_df.to_excel(writer, sheet_name='Параметры', index=False)
                worksheet = writer.sheets['Параметры']
                self._apply_styles(worksheet)
                self._auto_adjust_columns(worksheet)
                
                # ===============================================
                # ЛИСТ 5: Рекомендации (опционально)
                # ===============================================
                recommendations = []
                
                for i, section in enumerate(sections_data):
                    components = section.get("components", [])
                    
                    if components:
                        # Для сложных участков
                        for comp in components:
                            remaining = comp.get('remaining', 
                                                comp.get('thickness', 
                                                        comp.get('wall_thickness', 10)))
                            
                            if remaining < 4:
                                rec = f"{section['name']} - {comp.get('name', 'Компонент')}: АВАРИЙНОЕ состояние. Немедленный ремонт!"
                            elif remaining < 6:
                                rec = f"{section['name']} - {comp.get('name', 'Компонент')}: Плохое состояние. Ремонт в течение 3 месяцев."
                            elif remaining < 8:
                                rec = f"{section['name']} - {comp.get('name', 'Компонент')}: Удовлетворительно. Осмотр в течение 6 месяцев."
                            else:
                                rec = f"{section['name']} - {comp.get('name', 'Компонент')}: Нормальное состояние."
                            
                            recommendations.append(rec)
                    else:
                        # Для простых участков
                        remaining = section.get('remaining_thickness', section.get('thickness', 10))
                        
                        if remaining < 4:
                            rec = f"{section['name']}: АВАРИЙНОЕ состояние. Немедленный ремонт!"
                        elif remaining < 6:
                            rec = f"{section['name']}: Плохое состояние. Ремонт в течение 3 месяцев."
                        elif remaining < 8:
                            rec = f"{section['name']}: Удовлетворительно. Осмотр в течение 6 месяцев."
                        else:
                            rec = f"{section['name']}: Нормальное состояние."
                        
                        recommendations.append(rec)
                
                if recommendations:
                    rec_df = pd.DataFrame({'Рекомендации по ремонту и обслуживанию': recommendations})
                    rec_df.to_excel(writer, sheet_name='Рекомендации', index=False)
                    worksheet = writer.sheets['Рекомендации']
                    self._apply_styles(worksheet)
                    self._auto_adjust_columns(worksheet)
            
            return filename
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e

def create_export_dialog(parent, fluid_type, sections_data):
    """Диалог экспорта с выбором пути и имени файла"""
    
    dialog = tk.Toplevel(parent)
    dialog.title("Экспорт отчета в таблицу")
    dialog.geometry("600x700")
    dialog.resizable(False, False)
    dialog.transient(parent)
    dialog.grab_set()

    # ДОБАВЛЯЕМ ИКОНКУ ДИАЛОГУ
    icon_path = "assets/icon1.ico"
    if not os.path.exists(icon_path):
        icon_path = "../assets/icon1.ico"
    
    if os.path.exists(icon_path):
        try:
            dialog.iconbitmap(icon_path)
        except Exception as e:
            print(f"❌ Ошибка иконки диалога: {e}")
    
    # Центрируем
    dialog.geometry("+%d+%d" % (parent.winfo_rootx() + 100, parent.winfo_rooty() + 50))
    main_frame = ttk.Frame(dialog, padding=25)
    main_frame.pack(fill="both", expand=True)
    
    # Заголовок
    ttk.Label(main_frame, text="ЭКСПОРТ ОТЧЕТА В ТАБЛИЦУ", 
              font=("Arial", 16, "bold")).pack(pady=(0, 20))
    
    # Информация о проекте
    info_frame = ttk.LabelFrame(main_frame, text="ИНФОРМАЦИЯ О ПРОЕКТЕ", padding=10)
    info_frame.pack(fill="x", pady=(0, 15))
    
    # Подсчитываем компоненты
    total_components = 0
    for section in sections_data:
        components = section.get("components", [])
        if components:
            total_components += len(components)
        else:
            total_components += 1
    
    info_text = (
        f"Тип среды: {'НЕФТЬ' if fluid_type == 'oil' else 'ГАЗ'}\n"
        f"Количество участков: {len(sections_data)}\n"
        f"Количество компонентов: {total_components}\n"
        f"Дата: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    ttk.Label(info_frame, text=info_text, font=("Arial", 10)).pack(anchor="w")
    
    # Выбор имени файла
    name_frame = ttk.LabelFrame(main_frame, text="НАЗВАНИЕ ФАЙЛА", padding=10)
    name_frame.pack(fill="x", pady=(0, 15))
    
    # Переменная для имени файла
    filename_var = tk.StringVar()
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M')
    default_name = f"отчет_трубопровод_{timestamp}.xlsx"
    filename_var.set(default_name)
    
    ttk.Label(name_frame, text="Имя файла:", font=("Arial", 10)).pack(anchor="w")
    filename_entry = ttk.Entry(name_frame, textvariable=filename_var, width=50, font=("Arial", 10))
    filename_entry.pack(fill="x", pady=(5, 10))
    
    # Кнопка выбора папки
    def browse_folder():
        """Открывает диалог выбора папки"""
        folder = filedialog.askdirectory(
            title="Выберите папку для сохранения",
            initialdir=os.getcwd()
        )
        if folder:
            current_name = filename_var.get()
            # Если уже полный путь, берем только имя файла
            if os.path.dirname(current_name):
                current_name = os.path.basename(current_name)
            new_path = os.path.join(folder, current_name)
            filename_var.set(new_path)
    
    browse_btn = tk.Button(name_frame, text="Выбрать папку",
                          bg='#18171C', fg='#FADADD', font=("Arial", 9, "bold"),
                          relief='flat', borderwidth=0,
                          command=browse_folder)
    browse_btn.pack(pady=(0, 5))
    
    # Информация о содержимом
    content_frame = ttk.LabelFrame(main_frame, text="СОДЕРЖАНИЕ ОТЧЕТА", padding=10)
    content_frame.pack(fill="x", pady=(0, 20))
    
    content_text = (
        "Отчет будет содержать:\n"
        "• Лист 'Участки_сводка' - общая информация по участкам\n"
        "• Лист 'Компоненты' - детальная информация по всем компонентам\n"
        "• Лист 'Экономика' - сводка по стоимости ремонтов\n" 
        "• Лист 'Параметры' - технические параметры системы\n"
        "• Лист 'Рекомендации' - рекомендации по ремонту"
    )
    ttk.Label(content_frame, text=content_text, font=("Arial", 9), justify="left").pack(anchor="w")
    
    # Опции экспорта
    options_frame = ttk.LabelFrame(main_frame, text="НАСТРОЙКИ ЭКСПОРТА", padding=10)
    options_frame.pack(fill="x", pady=(0, 20))
    
    # Переменные для чекбоксов
    auto_open_var = tk.BooleanVar(value=True)
    detailed_export_var = tk.BooleanVar(value=True)
    
    ttk.Checkbutton(options_frame, text="Автоматически открыть файл после экспорта", 
                   variable=auto_open_var).pack(anchor="w", pady=2)
    ttk.Checkbutton(options_frame, text="Детализированный отчет (все компоненты)", 
                   variable=detailed_export_var, state="normal").pack(anchor="w", pady=2)
    
    # Кнопки управления
    button_frame = ttk.Frame(main_frame)
    button_frame.pack(fill="x", pady=10)
    
    def perform_export():
        """Выполняет экспорт"""
        filename = filename_var.get().strip()
        if not filename:
            messagebox.showerror("Ошибка", "Введите название файла!")
            return
        
        try:
            # Добавляем расширение если его нет
            if not filename.lower().endswith('.xlsx'):
                filename += '.xlsx'
            
            from models.economics import get_economic_summary
            economic_summary = get_economic_summary(sections_data)
            exporter = ReportExporter("Трубопровод", fluid_type)
            final_filename = exporter.export_to_excel(sections_data, economic_summary, filename)

            
            # Показываем успех
            if auto_open_var.get():
                result = messagebox.askyesno(
                    "Успех", 
                    f"Отчет успешно сохранен:\n{final_filename}\n\nОткрыть файл?",
                    icon='info'
                )
                
                if result:
                    # Пытаемся открыть файл
                    try:
                        os.startfile(final_filename)  
                    except:
                        try:
                            import subprocess
                            subprocess.run(['open', final_filename])  
                        except:
                            try:
                                subprocess.run(['xdg-open', final_filename])  
                            except:
                                messagebox.showinfo("Информация", f"Файл сохранен: {final_filename}")
            else:
                messagebox.showinfo(
                    "Успех", 
                    f"Отчет успешно сохранен:\n{final_filename}",
                    icon='info'
                )
            
            dialog.destroy()
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            
            messagebox.showerror("Ошибка", f"Не удалось экспортировать отчет:\n{type(e).__name__}: {e}")
    
    # Кнопка экспорта
    export_btn = tk.Button(button_frame, text="Экспортировать отчет",
                          bg='#18171C', fg='#FADADD', font=("Arial", 10, "bold"),
                          relief='flat', borderwidth=0,
                          command=perform_export)
    export_btn.pack(side="left", padx=(0, 10))
    
    # Кнопка предпросмотра
    def preview_data():
        """Предпросмотр данных для экспорта"""
        preview_dialog = tk.Toplevel(dialog)
        preview_dialog.title("Предпросмотр данных")
        preview_dialog.geometry("800x600")
        
        notebook = ttk.Notebook(preview_dialog)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Вкладка с участками
        sections_frame = ttk.Frame(notebook)
        sections_text = tk.Text(sections_frame, wrap="none")
        scrollbar_y = ttk.Scrollbar(sections_frame, orient="vertical", command=sections_text.yview)
        scrollbar_x = ttk.Scrollbar(sections_frame, orient="horizontal", command=sections_text.xview)
        sections_text.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        sections_text.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        
        # Добавляем данные
        sections_text.insert("1.0", "СВОДКА ПО УЧАСТКАМ:\n" + "="*50 + "\n\n")
        for i, section in enumerate(sections_data, 1):
            sections_text.insert("end", f"{i}. {section['name']}\n")
            components = section.get("components", [])
            if components:
                sections_text.insert("end", f"   Компонентов: {len(components)}\n")
                for comp in components:
                    sections_text.insert("end", f"   - {comp.get('name', 'Компонент')}: {comp.get('component_type', 'unknown')}\n")
            sections_text.insert("end", "\n")
        
        notebook.add(sections_frame, text="Участки")
        
        # Вкладка с экономикой
        try:
            from models.economics import get_economic_summary
            econ_frame = ttk.Frame(notebook)
            econ_text = tk.Text(econ_frame, wrap="word")
            econ_scrollbar = ttk.Scrollbar(econ_frame, orient="vertical", command=econ_text.yview)
            econ_text.configure(yscrollcommand=econ_scrollbar.set)
            
            econ_text.pack(side="left", fill="both", expand=True)
            econ_scrollbar.pack(side="right", fill="y")
            
            economic_summary = get_economic_summary(sections_data)
            econ_text.insert("1.0", "ЭКОНОМИЧЕСКАЯ СВОДКА:\n" + "="*50 + "\n\n")
            for key, value in economic_summary.items():
                if isinstance(value, (int, float)):
                    econ_text.insert("end", f"{key}: {value:,.0f}\n")
                else:
                    econ_text.insert("end", f"{key}: {value}\n")
            
            notebook.add(econ_frame, text="Экономика")
        except Exception as e:
            print(f"Ошибка загрузки экономики: {e}")
    
    preview_btn = tk.Button(button_frame, text="Предпросмотр",
                           bg='#18171C', fg='#FADADD', font=("Arial", 10, "bold"),
                           relief='flat', borderwidth=0,
                           command=preview_data)
    preview_btn.pack(side="left")
    
    # Кнопка отмены
    cancel_btn = tk.Button(button_frame, text="Отмена", 
                          bg='#18171C', fg='#FADADD', font=("Arial", 10, "bold"),
                          relief='flat', borderwidth=0,
                          command=dialog.destroy)
    cancel_btn.pack(side="right")
    
    # Фокус на поле ввода
    filename_entry.focus()
    filename_entry.select_range(0, tk.END)
    
    # Привязываем Enter к экспорту
    def on_enter(event):
        perform_export()
    
    dialog.bind('<Return>', on_enter)
